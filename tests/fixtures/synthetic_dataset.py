"""Synthetic ParcelPilot-shaped fixtures for ingestion/rule-engine/etc. tests.

These mirror the *shape* of the real assessment data pack — a current vs.
deprecated policy pair, a customer contract that waives a fee and tightens
SLA, a second contract that overrides a service-credit threshold/amount,
multiple accounts/orders/tickets, and a historical ticket resolution that
contradicts an active contract — using entirely fictional companies and
made-up numbers.

The test suite depends on this module, never on the real proprietary pack,
which is itself the proof that the system generalizes rather than being
hard-coded to the assessment brief's specific example records (see
docs/git-development-plan.md §2).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import yaml

# ---------------------------------------------------------------------------
# Structured workbook fixture
# ---------------------------------------------------------------------------


def write_synthetic_workbook(path: Path) -> None:
    """Write a workbook with the same sheet/column shape as the real pack."""
    workbook = openpyxl.Workbook()

    readme = workbook.active
    readme.title = "README"
    readme.append(["ParcelPilot AI Agent Assessment - Structured Data (fixture)", None])
    readme.append(["Dataset snapshot", "2027-01-10 09:00 Asia/Kolkata"])
    readme.append(["Currency", "INR"])
    readme.append(["Notes", "Synthetic fixture dataset for automated tests."])
    readme.append(
        ["Important", "Historical ticket resolutions are context only and may be incorrect."]
    )

    accounts = workbook.create_sheet("accounts")
    accounts.append(
        [
            "account_id",
            "account_name",
            "plan",
            "status",
            "csm",
            "contract_file",
            "premium_support",
            "notes",
        ]
    )
    accounts.append(
        [
            "FIX-ACCT-001",
            "Fixture Freight Co",
            "Enterprise",
            "active",
            "Fixture CSM A",
            "fixture_contract_a.txt",
            True,
            "Has a contract waiving the cancellation fee and tightening SLA.",
        ]
    )
    accounts.append(
        [
            "FIX-ACCT-002",
            "Fixture Cargo Ltd",
            "Growth",
            "active",
            "Fixture CSM B",
            "fixture_contract_b.txt",
            False,
            "Has a contract overriding the service-credit threshold/amount.",
        ]
    )
    accounts.append(
        [
            "FIX-ACCT-003",
            "Fixture Retail Inc",
            "Standard",
            "active",
            "Fixture CSM C",
            None,
            False,
            "No custom agreement; standard policy applies.",
        ]
    )

    orders = workbook.create_sheet("orders")
    orders.append(
        [
            "order_id",
            "account_id",
            "carrier",
            "status",
            "booked_at",
            "pickup_window_start",
            "pickup_window_end",
            "pickup_actual_at",
            "shipment_fee_inr",
            "carrier_fault",
            "customer_fault",
            "cancellation_requested_at",
            "notes",
        ]
    )
    orders.append(
        [
            "FIX-ORD-001",
            "FIX-ACCT-001",
            "FixtureCarrier",
            "BOOKED",
            "2027-01-10 09:00",
            "2027-01-10 10:00",
            "2027-01-10 11:00",
            None,
            1000.0,
            False,
            False,
            "2027-01-10 11:30",
            "Cancellation requested 2.5 hours after booking; not yet picked up.",
        ]
    )
    orders.append(
        [
            "FIX-ORD-002",
            "FIX-ACCT-002",
            "FixtureCarrier",
            "BOOKED",
            "2027-01-10 03:00",
            "2027-01-10 04:00",
            "2027-01-10 05:00",
            None,
            2000.0,
            True,
            False,
            None,
            "Pickup missed, carrier at fault, still not picked up at snapshot.",
        ]
    )
    orders.append(
        [
            "FIX-ORD-003",
            "FIX-ACCT-003",
            "FixtureCarrier",
            "DELIVERED",
            "2027-01-08 09:00",
            "2027-01-08 10:00",
            "2027-01-08 11:00",
            "2027-01-08 10:20",
            1500.0,
            False,
            False,
            None,
            "Completed delivery.",
        ]
    )

    tickets = workbook.create_sheet("tickets")
    tickets.append(
        [
            "ticket_id",
            "account_id",
            "created_at",
            "status",
            "subject",
            "description",
            "channel",
            "assigned_to",
            "last_customer_message_at",
            "historical_resolution",
        ]
    )
    tickets.append(
        [
            "FIX-TKT-001",
            "FIX-ACCT-001",
            "2027-01-10 08:30",
            "open",
            "All shipment creation is failing",
            "Every user gets HTTP 500 when creating any shipment.",
            "email",
            "Fixture Agent A",
            "2027-01-10 08:45",
            None,
        ]
    )
    tickets.append(
        [
            "FIX-TKT-002",
            "FIX-ACCT-002",
            "2026-12-01 10:00",
            "closed",
            "Cancellation fee question",
            "Asked whether a fee applied 90 minutes after booking.",
            "chat",
            "Fixture Agent B",
            "2026-12-01 10:30",
            "Agent told the customer a fee applied — incorrect, given the account's contract.",
        ]
    )

    workbook.save(path)


# ---------------------------------------------------------------------------
# Document corpus fixture
# ---------------------------------------------------------------------------

_DOC_MANIFEST: dict[str, list[dict[str, object]]] = {
    "documents": [
        {
            "filename": "fixture_policy_v2_current.txt",
            "document_type": "policy",
            "version": "v2",
            "status": "current",
            "effective_date": "2027-01-01",
            "customer_account_id": None,
            "authority_tier": "2",
        },
        {
            "filename": "fixture_policy_v1_deprecated.txt",
            "document_type": "policy",
            "version": "v1",
            "status": "deprecated",
            "effective_date": "2026-01-01",
            "superseded_by": "fixture_policy_v2_current.txt",
            "customer_account_id": None,
            "authority_tier": "excluded",
        },
        {
            "filename": "fixture_sop.txt",
            "document_type": "sop",
            "version": "v1",
            "status": "current",
            "effective_date": "2027-01-01",
            "customer_account_id": None,
            "authority_tier": "2",
        },
        {
            "filename": "fixture_contract_a.txt",
            "document_type": "contract",
            "version": None,
            "status": "active",
            "effective_date": "2027-01-01",
            "customer_account_id": "FIX-ACCT-001",
            "authority_tier": "1",
        },
        {
            "filename": "fixture_contract_b.txt",
            "document_type": "contract",
            "version": None,
            "status": "active",
            "effective_date": "2027-01-01",
            "customer_account_id": "FIX-ACCT-002",
            "authority_tier": "1",
        },
    ]
}

_DOC_TEXTS: dict[str, str] = {
    "fixture_policy_v2_current.txt": (
        "Fixture Support Policy v2\n"
        "Status: CURRENT\n"
        "Effective: 2027-01-01\n"
        "1. Severity definitions\n"
        "P1 is a complete outage. P2 is a major degradation. P3 is a minor issue.\n"
        "2. Default first-response targets\n"
        "Enterprise P1 target is 1 hour. Growth P1 target is 4 hours.\n"
    ),
    "fixture_policy_v1_deprecated.txt": (
        "Fixture Support Policy v1\n"
        "Status: DEPRECATED - DO NOT USE FOR CURRENT REQUESTS\n"
        "Effective: 2026-01-01\n"
        "Enterprise P1 target is 8 hours. This value is superseded and must not be used.\n"
    ),
    "fixture_sop.txt": (
        "Fixture Cancellation and Service Credit SOP\n"
        "Status: CURRENT\n"
        "1. Order cancellation\n"
        "BOOKED orders may be cancelled free of charge within 30 minutes of booking; a fee "
        "applies after that unless a customer agreement waives it.\n"
        "2. Failed-pickup service credits\n"
        "A customer is eligible for a credit when the pickup is more than 2 hours late due to "
        "carrier fault. The default credit is the lower of 500 or 10 percent of the shipment fee.\n"
    ),
    "fixture_contract_a.txt": (
        "Fixture Freight Co Agreement\n"
        "Account: FIX-ACCT-001\n"
        "Status: ACTIVE\n"
        "1. Shipment cancellation\n"
        "Fixture Freight Co may cancel any BOOKED shipment with no cancellation fee, regardless "
        "of how long ago it was booked.\n"
        "2. Support terms\n"
        "P1 target is 15 minutes, 24x7.\n"
    ),
    "fixture_contract_b.txt": (
        "Fixture Cargo Ltd Agreement\n"
        "Account: FIX-ACCT-002\n"
        "Status: ACTIVE\n"
        "1. Failed-pickup credits\n"
        "If a pickup is more than 4 hours late due to carrier fault, Fixture Cargo Ltd receives a "
        "fixed 300 credit. This replaces the default threshold and amount.\n"
    ),
}


def write_synthetic_documents(documents_dir: Path, manifest_path: Path) -> None:
    """Write the fixture document corpus (plain-text) and its manifest."""
    documents_dir.mkdir(parents=True, exist_ok=True)
    for filename, text in _DOC_TEXTS.items():
        (documents_dir / filename).write_text(text, encoding="utf-8")
    manifest_path.write_text(yaml.safe_dump(_DOC_MANIFEST, sort_keys=False), encoding="utf-8")

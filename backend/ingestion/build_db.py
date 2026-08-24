"""Idempotent ingestion of the ParcelPilot structured-data workbook into SQLite.

Reads the workbook described in `data/README.md` (`README`, `accounts`,
`orders`, `tickets` sheets) and rebuilds the corresponding SQLite tables.
Safe to re-run at any time: each run fully replaces prior contents, so the
database always reflects exactly what is currently in the source workbook,
with no leftover rows from a previous version or a previously-supplied
(and since-swapped) data pack.

The dataset's reference "now" (see docs/system-design.md, Part Q) is read
from the workbook's README sheet, not the system clock, and every naive
timestamp in `orders`/`tickets` is normalized to that same timezone and
stored as a timezone-aware ISO-8601 string, so downstream code never has to
guess what timezone a bare "YYYY-MM-DD HH:MM" string is in.

Usage:
    python -m backend.ingestion.build_db [--workbook PATH] [--db-path PATH]
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, tzinfo
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from backend.config import settings

_SNAPSHOT_PATTERN = re.compile(
    r"^(?P<naive>\d{4}-\d{2}-\d{2} \d{2}:\d{2})\s+(?P<tz>[A-Za-z0-9_/+\-]+)$"
)
_TIMESTAMP_FORMAT = "%Y-%m-%d %H:%M"

_SCHEMA = """
DROP TABLE IF EXISTS accounts;
DROP TABLE IF EXISTS orders;
DROP TABLE IF EXISTS tickets;
DROP TABLE IF EXISTS dataset_meta;

CREATE TABLE dataset_meta (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE accounts (
    account_id TEXT PRIMARY KEY,
    account_name TEXT NOT NULL,
    plan TEXT NOT NULL,
    status TEXT NOT NULL,
    csm TEXT,
    contract_file TEXT,
    premium_support INTEGER NOT NULL DEFAULT 0,
    notes TEXT
);

CREATE TABLE orders (
    order_id TEXT PRIMARY KEY,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    carrier TEXT,
    status TEXT NOT NULL,
    booked_at TEXT NOT NULL,
    pickup_window_start TEXT,
    pickup_window_end TEXT,
    pickup_actual_at TEXT,
    shipment_fee_inr REAL,
    carrier_fault INTEGER NOT NULL DEFAULT 0,
    customer_fault INTEGER NOT NULL DEFAULT 0,
    cancellation_requested_at TEXT,
    notes TEXT
);

CREATE TABLE tickets (
    ticket_id TEXT PRIMARY KEY,
    account_id TEXT NOT NULL REFERENCES accounts(account_id),
    created_at TEXT NOT NULL,
    status TEXT NOT NULL,
    subject TEXT,
    description TEXT,
    channel TEXT,
    assigned_to TEXT,
    last_customer_message_at TEXT,
    historical_resolution TEXT
);
"""


@dataclass(frozen=True)
class DatasetMeta:
    snapshot_time: datetime  # timezone-aware
    currency: str
    notes: str
    important: str


def parse_snapshot(value: str) -> datetime:
    """Parse README snapshot strings such as "2026-08-16 11:00 Asia/Kolkata".

    Raises ValueError if the value doesn't match the expected
    "<date> <time> <IANA timezone name>" shape, since silently falling back
    to some default timezone would defeat the point of sourcing "now" from
    the dataset rather than assuming one.
    """
    match = _SNAPSHOT_PATTERN.match(value.strip())
    if not match:
        raise ValueError(f"Unrecognized dataset snapshot format: {value!r}")
    tz = ZoneInfo(match.group("tz"))
    naive = datetime.strptime(match.group("naive"), _TIMESTAMP_FORMAT)
    return naive.replace(tzinfo=tz)


def parse_local_timestamp(value: object, tz: tzinfo) -> str | None:
    """Attach the dataset's timezone to a naive "YYYY-MM-DD HH:MM" cell value
    and return it as an ISO-8601 string, or None if the cell is empty.
    """
    if value is None or value == "":
        return None
    naive = datetime.strptime(str(value).strip(), _TIMESTAMP_FORMAT)
    return naive.replace(tzinfo=tz).isoformat()


def read_readme(ws: Worksheet) -> DatasetMeta:
    values: dict[str, object] = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            values[str(row[0])] = row[1] if len(row) > 1 else None

    if "Dataset snapshot" not in values:
        raise ValueError("README sheet is missing a 'Dataset snapshot' row")

    return DatasetMeta(
        snapshot_time=parse_snapshot(str(values["Dataset snapshot"])),
        currency=str(values.get("Currency", "")),
        notes=str(values.get("Notes", "")),
        important=str(values.get("Important", "")),
    )


def read_sheet_rows(ws: Worksheet) -> list[dict[str, object]]:
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(cell) for cell in next(rows_iter)]
    return [
        dict(zip(header, row, strict=True))
        for row in rows_iter
        if any(cell is not None for cell in row)
    ]


def _insert_accounts(conn: sqlite3.Connection, rows: list[dict[str, object]]) -> None:
    conn.executemany(
        """INSERT INTO accounts
           (account_id, account_name, plan, status, csm, contract_file, premium_support, notes)
           VALUES (:account_id, :account_name, :plan, :status, :csm, :contract_file,
                   :premium_support, :notes)""",
        [
            {
                "account_id": row["account_id"],
                "account_name": row["account_name"],
                "plan": row["plan"],
                "status": row["status"],
                "csm": row.get("csm"),
                "contract_file": row.get("contract_file"),
                "premium_support": 1 if row.get("premium_support") else 0,
                "notes": row.get("notes"),
            }
            for row in rows
        ],
    )


def _insert_orders(conn: sqlite3.Connection, rows: list[dict[str, object]], tz: tzinfo) -> None:
    conn.executemany(
        """INSERT INTO orders
           (order_id, account_id, carrier, status, booked_at, pickup_window_start,
            pickup_window_end, pickup_actual_at, shipment_fee_inr, carrier_fault,
            customer_fault, cancellation_requested_at, notes)
           VALUES (:order_id, :account_id, :carrier, :status, :booked_at,
                   :pickup_window_start, :pickup_window_end, :pickup_actual_at,
                   :shipment_fee_inr, :carrier_fault, :customer_fault,
                   :cancellation_requested_at, :notes)""",
        [
            {
                "order_id": row["order_id"],
                "account_id": row["account_id"],
                "carrier": row.get("carrier"),
                "status": row["status"],
                "booked_at": parse_local_timestamp(row.get("booked_at"), tz),
                "pickup_window_start": parse_local_timestamp(row.get("pickup_window_start"), tz),
                "pickup_window_end": parse_local_timestamp(row.get("pickup_window_end"), tz),
                "pickup_actual_at": parse_local_timestamp(row.get("pickup_actual_at"), tz),
                "shipment_fee_inr": row.get("shipment_fee_inr"),
                "carrier_fault": 1 if row.get("carrier_fault") else 0,
                "customer_fault": 1 if row.get("customer_fault") else 0,
                "cancellation_requested_at": parse_local_timestamp(
                    row.get("cancellation_requested_at"), tz
                ),
                "notes": row.get("notes"),
            }
            for row in rows
        ],
    )


def _insert_tickets(conn: sqlite3.Connection, rows: list[dict[str, object]], tz: tzinfo) -> None:
    conn.executemany(
        """INSERT INTO tickets
           (ticket_id, account_id, created_at, status, subject, description,
            channel, assigned_to, last_customer_message_at, historical_resolution)
           VALUES (:ticket_id, :account_id, :created_at, :status, :subject, :description,
                   :channel, :assigned_to, :last_customer_message_at, :historical_resolution)""",
        [
            {
                "ticket_id": row["ticket_id"],
                "account_id": row["account_id"],
                "created_at": parse_local_timestamp(row.get("created_at"), tz),
                "status": row["status"],
                "subject": row.get("subject"),
                "description": row.get("description"),
                "channel": row.get("channel"),
                "assigned_to": row.get("assigned_to"),
                "last_customer_message_at": parse_local_timestamp(
                    row.get("last_customer_message_at"), tz
                ),
                "historical_resolution": row.get("historical_resolution"),
            }
            for row in rows
        ],
    )


def build_database(workbook_path: Path, db_path: Path) -> DatasetMeta:
    """Rebuild `db_path` from `workbook_path`. Returns the parsed dataset metadata."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    meta = read_readme(workbook["README"])
    tz = meta.snapshot_time.tzinfo
    assert tz is not None  # parse_snapshot always attaches a tzinfo

    accounts = read_sheet_rows(workbook["accounts"])
    orders = read_sheet_rows(workbook["orders"])
    tickets = read_sheet_rows(workbook["tickets"])

    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(_SCHEMA)

        conn.executemany(
            "INSERT INTO dataset_meta (key, value) VALUES (?, ?)",
            [
                ("snapshot_time", meta.snapshot_time.isoformat()),
                ("currency", meta.currency),
                ("notes", meta.notes),
                ("important", meta.important),
                ("source_workbook", workbook_path.name),
            ],
        )

        _insert_accounts(conn, accounts)
        _insert_orders(conn, orders, tz)
        _insert_tickets(conn, tickets, tz)

        conn.commit()
    finally:
        conn.close()

    return meta


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=settings.data_dir / "ParcelPilot_Assessment_Data.xlsx",
        help="Path to the ParcelPilot assessment workbook.",
    )
    parser.add_argument(
        "--db-path",
        type=Path,
        default=settings.var_dir / "app.db",
        help="Path to the SQLite database to (re)build.",
    )
    args = parser.parse_args()

    meta = build_database(args.workbook, args.db_path)
    print(f"Ingested {args.workbook} -> {args.db_path}")
    print(f"  dataset snapshot: {meta.snapshot_time.isoformat()}")


if __name__ == "__main__":
    main()
